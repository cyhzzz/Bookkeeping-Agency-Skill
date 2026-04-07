"""
畅捷通T+/T6标准导出格式解析器
支持版本: T+13.0, T6v6.1
导出格式: .xlsx (Excel格式)
"""

import openpyxl
from pathlib import Path
from typing import List, Optional
from datetime import datetime

from .base_parser import (
    BaseParser,
    ParsedFinancialData,
    ParsedVoucher,
    VoucherEntry
)


class ChanjetParser(BaseParser):
    """畅捷通解析器"""

    @property
    def software_name(self) -> str:
        return "畅捷通T+/T6"

    @property
    def supported_formats(self) -> List[str]:
        return [".xlsx"]

    def can_parse(self, file_path: str) -> bool:
        """判断是否能解析此文件"""
        path = Path(file_path)
        if path.suffix.lower() not in self.supported_formats:
            return False

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            # 畅捷通特征工作表名
            return any('凭证库' in s or '科目余额' in s for s in sheet_names)
        except Exception:
            return False

    def parse(self, file_path: str) -> ParsedFinancialData:
        """解析畅捷通导出的Excel文件"""
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # 查找凭证库工作表
        voucher_sheet = None
        for sheet_name in wb.sheetnames:
            if '凭证库' in sheet_name or '凭证' in sheet_name:
                voucher_sheet = wb[sheet_name]
                break

        vouchers = []
        if voucher_sheet:
            vouchers = self._parse_voucher_sheet(voucher_sheet, file_path)

        company_name = Path(file_path).stem  # 使用文件名作为公司名

        return ParsedFinancialData(
            parser_name=self.software_name,
            company_name=company_name,
            period=datetime.now().strftime("%Y-%m"),
            vouchers=vouchers,
            raw_data={'source_file': file_path},
            confidence=0.85
        )

    def _parse_voucher_sheet(self, sheet, source_file: str) -> List[ParsedVoucher]:
        """解析凭证库工作表"""
        vouchers = []

        try:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []

            header = [str(h).strip() if h else "" for h in rows[0]]
            col_map = {h: i for i, h in enumerate(header)}

            current_voucher = None
            current_entries = []
            voucher_id = 1

            for row in rows[1:]:
                if not row or not any(row):
                    continue

                cells = [str(c).strip() if c else "" for c in row]

                # 判断行类型
                if self._is_new_voucher(cells, col_map):
                    # 保存上一个凭证
                    if current_voucher and current_entries:
                        current_voucher.entries = current_entries
                        vouchers.append(current_voucher)

                    # 新建凭证
                    current_voucher = ParsedVoucher(
                        voucher_id=f"V{voucher_id:04d}",
                        date=self._get_cell_value(cells, col_map, '凭证日期', '日期'),
                        voucher_no=self._get_cell_value(cells, col_map, '凭证编号', '凭证号'),
                        summary=self._get_cell_value(cells, col_map, '摘要'),
                        source_file=source_file
                    )
                    current_entries = []
                    voucher_id += 1

                elif self._is_entry_row(cells) and current_voucher:
                    entry = self._parse_entry(cells, col_map)
                    if entry:
                        current_entries.append(entry)

            # 保存最后一个凭证
            if current_voucher and current_entries:
                current_voucher.entries = current_entries
                vouchers.append(current_voucher)

        except Exception as e:
            print(f"解析凭证库出错: {e}")

        return vouchers

    def _is_new_voucher(self, cells: List[str], col_map: dict) -> bool:
        """判断是否为新凭证行"""
        date_val = self._get_cell_value(cells, col_map, '凭证日期', '日期')
        if date_val and self._looks_like_date(date_val):
            return True
        return False

    def _looks_like_date(self, value: str) -> bool:
        """判断是否像日期"""
        import re
        return bool(re.match(r'\d{4}[-/]\d{1,2}', value))

    def _is_entry_row(self, cells: List[str]) -> bool:
        """判断是否为分录行"""
        try:
            # 畅捷通通常有借方贷方两列
            for c in cells:
                if c:
                    try:
                        float(c.replace(',', '').replace(' ', ''))
                        return True
                    except ValueError:
                        continue
        except Exception:
            pass
        return False

    def _parse_entry(self, cells: List[str], col_map: dict) -> Optional[VoucherEntry]:
        """解析分录"""
        account_name = self._get_cell_value(cells, col_map, '科目名称', '科目')
        if not account_name:
            return None

        account_code = self._get_cell_value(cells, col_map, '科目编码')

        debit_str = self._get_cell_value(cells, col_map, '借方')
        credit_str = self._get_cell_value(cells, col_map, '贷方')

        try:
            debit = float(debit_str.replace(',', '').replace(' ', '') or 0) if debit_str else 0
            credit = float(credit_str.replace(',', '').replace(' ', '') or 0) if credit_str else 0
        except ValueError:
            debit, credit = 0, 0

        return VoucherEntry(
            account_code=account_code,
            account_name=account_name,
            debit_amount=debit,
            credit_amount=credit
        )

    def _get_cell_value(self, cells: List[str], col_map: dict, *keys) -> str:
        """获取单元格值"""
        for key in keys:
            if key in col_map and col_map[key] < len(cells):
                val = cells[col_map[key]]
                if val and val != 'None':
                    return val
        return ""
