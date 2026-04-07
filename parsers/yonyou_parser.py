"""
用友U8/U9/NC标准导出格式解析器
支持版本: U8v13.0, U9v5.0, NC6.5
导出格式: .xlsx (Excel格式)
"""

import openpyxl
from pathlib import Path
from typing import List, Optional
from datetime import datetime

from .base_parser import (
    BaseParser,
    ParsedFinancialData,
    ParsedVoucher,
    VoucherEntry
)


class YonyouParser(BaseParser):
    """用友解析器"""

    @property
    def software_name(self) -> str:
        return "用友U8/U9"

    @property
    def supported_formats(self) -> List[str]:
        return [".xlsx", ".xls"]

    def can_parse(self, file_path: str) -> bool:
        """判断是否能解析此文件"""
        path = Path(file_path)
        if path.suffix.lower() not in self.supported_formats:
            return False

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = [s.lower() for s in wb.sheetnames]
            # 用友特征工作表名
            return any(k in s for s in sheet_names
                      for k in ['记账凭证', '明细账', '用友'])
        except Exception:
            return False

    def parse(self, file_path: str) -> ParsedFinancialData:
        """解析用友导出的Excel文件"""
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # 查找凭证工作表
        voucher_sheet = None
        for sheet_name in wb.sheetnames:
            if '记账凭证' in sheet_name or '凭证' in sheet_name:
                voucher_sheet = wb[sheet_name]
                break

        if voucher_sheet:
            vouchers = self._parse_voucher_sheet(voucher_sheet, file_path)
        else:
            vouchers = []

        # 提取公司名称（从工作簿属性）
        company_name = getattr(wb.properties, 'title', None) or "未知公司"

        return ParsedFinancialData(
            parser_name=self.software_name,
            company_name=str(company_name),
            period=datetime.now().strftime("%Y-%m"),
            vouchers=vouchers,
            raw_data={'source_file': file_path},
            confidence=0.85
        )

    def _parse_voucher_sheet(self, sheet, source_file: str) -> List[ParsedVoucher]:
        """解析凭证工作表"""
        vouchers = []

        try:
            # 读取所有行
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []

            # 第一行是表头
            header = [str(h).strip() if h else "" for h in rows[0]]

            # 建立列索引映射
            col_map = {h: i for i, h in enumerate(header)}

            current_voucher = None
            current_entries = []
            voucher_id = 1

            for row in rows[1:]:
                if not row or not any(row):
                    continue

                cells = [str(c).strip() if c else "" for c in row]

                # 判断行类型
                if self._is_voucher_header_row(cells, col_map):
                    # 保存上一个凭证
                    if current_voucher and current_entries:
                        current_voucher.entries = current_entries
                        vouchers.append(current_voucher)

                    # 新建凭证
                    current_voucher = ParsedVoucher(
                        voucher_id=f"V{voucher_id:04d}",
                        date=self._get_cell_value(cells, col_map, '日期', '制单日期'),
                        voucher_no=self._get_cell_value(cells, col_map, '凭证号', '凭证字号'),
                        summary=self._get_cell_value(cells, col_map, '摘要'),
                        source_file=source_file
                    )
                    current_entries = []
                    voucher_id += 1

                elif self._is_entry_row(cells, col_map) and current_voucher:
                    # 解析分录
                    entry = self._parse_entry(cells, col_map)
                    if entry:
                        current_entries.append(entry)

            # 保存最后一个凭证
            if current_voucher and current_entries:
                current_voucher.entries = current_entries
                vouchers.append(current_voucher)

        except Exception as e:
            print(f"解析凭证工作表出错: {e}")

        return vouchers

    def _is_voucher_header_row(self, cells: List[str], col_map: dict) -> bool:
        """判断是否为凭证表头行"""
        # 检查日期列是否有日期格式
        date_val = self._get_first_matching_cell(cells, col_map, ['日期', '制单日期'])
        if date_val and self._looks_like_date(date_val):
            return True
        return False

    def _looks_like_date(self, value: str) -> bool:
        """判断是否像日期"""
        import re
        return bool(re.match(r'\d{4}[-/]\d{1,2}', value))

    def _is_entry_row(self, cells: List[str], col_map: dict) -> bool:
        """判断是否为分录行"""
        # 检查是否有金额列且有数值
        amount_val = self._get_first_matching_cell(cells, col_map, ['借方金额', '贷方金额', '金额', '借方', '贷方'])
        if amount_val:
            try:
                float(amount_val.replace(',', '').replace(' ', ''))
                return True
            except ValueError:
                pass
        return False

    def _parse_entry(self, cells: List[str], col_map: dict) -> Optional[VoucherEntry]:
        """解析分录"""
        account_name = self._get_first_matching_cell(cells, col_map,
            ['科目', '会计科目', '科目名称', '科目编码'])
        if not account_name:
            return None

        debit_str = self._get_first_matching_cell(cells, col_map,
            ['借方金额', '借方发生', '借方'])
        credit_str = self._get_first_matching_cell(cells, col_map,
            ['贷方金额', '贷方发生', '贷方'])

        try:
            debit = float(debit_str.replace(',', '').replace(' ', '') or 0) if debit_str else 0
            credit = float(credit_str.replace(',', '').replace(' ', '') or 0) if credit_str else 0
        except ValueError:
            debit, credit = 0, 0

        return VoucherEntry(
            account_code="",
            account_name=account_name,
            debit_amount=debit,
            credit_amount=credit
        )

    def _get_cell_value(self, cells: List[str], col_map: dict, *keys) -> str:
        """获取单元格值"""
        val = self._get_first_matching_cell(cells, col_map, keys)
        return val if val else ""

    def _get_first_matching_cell(self, cells: List[str], col_map: dict, keys: List[str]) -> str:
        """获取第一个匹配的关键字对应的单元格值"""
        for key in keys:
            if key in col_map and col_map[key] < len(cells):
                return cells[col_map[key]]
        return ""
